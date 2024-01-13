import pandas as pd
import openpyxl

def process_sheet(sheet):
    # Find merged cells in rows 2 and 3
    merged_cells_ranges = sheet.merged_cells.ranges
    merged_headers = {}
    for merged_range in merged_cells_ranges:
        if merged_range.min_row <= 3 and merged_range.max_row >= 2:
            top_left_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                header_cell = sheet.cell(row=4, column=col).value
                merged_headers[col] = f"{top_left_cell} {header_cell}" if header_cell else top_left_cell

    # Create a DataFrame from the sheet, skipping the first 3 rows
    df = pd.DataFrame(sheet.values)
    df = df.iloc[3:]  # Skip the first 3 rows

    # Replace column names with merged headers where applicable
    for col_index, header in merged_headers.items():
        df.iloc[0, col_index - 1] = header

    # Set the first row as header
    new_header = df.iloc[0]
    df = df[1:]
    df.columns = new_header

    return df

def convert_excel_to_csv(file_path):
    workbook = openpyxl.load_workbook(file_path)

    for sheet_name in workbook.sheetnames[1:]: # Skip the first sheet
        df = process_sheet(workbook[sheet_name])
        csv_file_path = f'{sheet_name}.csv'
        df.to_csv(csv_file_path, index=False)
        print(f"Sheet '{sheet_name}' has been saved as '{csv_file_path}'")

if __name__ == "__main__":
    file_path = 'bba_06_2023.xlsx'  # Replace with your file path
    convert_excel_to_csv(file_path)